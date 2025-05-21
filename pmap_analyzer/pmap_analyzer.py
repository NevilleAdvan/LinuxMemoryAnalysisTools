import sys
import argparse
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def parse_pmap_output(pmap_content):
    """解析pmap输出并按Mapping类型统计内存使用情况"""
    stats = defaultdict(lambda: {'Kbytes': 0, 'PSS': 0, 'Dirty': 0})
    parsing_data = False

    for line in pmap_content:
        line = line.strip()

        if line.startswith('Address') or line.startswith('------'):
            parsing_data = True
            continue

        if not parsing_data or line.startswith('total') or line.startswith('---'):
            continue

        parts = line.split()
        if len(parts) < 7:  # 至少需要7列（Address, Kbytes, PSS, Dirty, Swap, Mode, Mapping）
            continue

        try:
            kbytes = int(parts[1])
            pss = int(parts[2])
            dirty = int(parts[3])
            # Swap列在parts[4]，Mode列在parts[5]
            mode = parts[5]  # 修正此处！之前错误使用parts[4]
            mapping_parts = parts[6:]  # Mapping从第7列开始（索引6）
            mapping = ' '.join(mapping_parts) if mapping_parts else 'Unknown'

            if mapping.startswith('[') and mapping.endswith(']'):
                mapping = mapping.strip('[]')

            if not mapping:
                mapping = 'Unknown'

        except (ValueError, IndexError):
            continue

        key = (mode, mapping)
        stats[key]['Kbytes'] += kbytes
        stats[key]['PSS'] += pss
        stats[key]['Dirty'] += dirty

    return stats


def format_output(stats):
    """格式化输出结果，确保各列对齐"""
    max_mode_width = 7
    max_mapping_width = 8
    max_kbytes_width = 5
    max_pss_width = 3
    max_dirty_width = 5

    for (mode, mapping) in stats:
        max_mode_width = max(max_mode_width, len(mode))
        max_mapping_width = max(max_mapping_width, len(mapping))
        max_kbytes_width = max(max_kbytes_width, len(str(stats[(mode, mapping)]['Kbytes'])))
        max_pss_width = max(max_pss_width, len(str(stats[(mode, mapping)]['PSS'])))
        max_dirty_width = max(max_dirty_width, len(str(stats[(mode, mapping)]['Dirty'])))

    header_format = f"{{:<{max_mode_width}}} {{:<{max_mapping_width}}} {{:>{max_kbytes_width}}} {{:>{max_pss_width}}} {{:>{max_dirty_width}}}"
    row_format = f"{{:<{max_mode_width}}} {{:<{max_mapping_width}}} {{:>{max_kbytes_width}d}} {{:>{max_pss_width}d}} {{:>{max_dirty_width}d}}"

    header = header_format.format("Mode", "Mapping", "Kbytes", "PSS", "Dirty")
    separator = '-' * len(header)

    rows = []
    for (mode, mapping) in sorted(stats, key=lambda k: stats[k]['Kbytes'], reverse=True):
        data = stats[(mode, mapping)]
        rows.append(row_format.format(mode, mapping, data['Kbytes'], data['PSS'], data['Dirty']))

    return [header, separator] + rows


def write_to_excel(stats, filename):
    """将统计结果写入Excel文件"""
    # 创建工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = "内存映射统计"

    # 添加表头
    headers = ["Mode", "Mapping", "Kbytes", "PSS", "Dirty"]
    ws.append(headers)

    # 设置表头样式
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # 添加数据行
    for (mode, mapping) in sorted(stats, key=lambda k: stats[k]['Kbytes'], reverse=True):
        data = stats[(mode, mapping)]
        ws.append([mode, mapping, data['Kbytes'], data['PSS'], data['Dirty']])

    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # 保存Excel文件
    try:
        wb.save(filename)
        print(f"已将结果保存到Excel文件: {filename}")
    except Exception as e:
        print(f"保存Excel文件时出错: {e}")


def main():
    """主函数：处理输入并输出统计结果"""
    # 创建参数解析器
    parser = argparse.ArgumentParser(description='分析pmap输出并统计内存使用情况')
    parser.add_argument('-i', '--input', help='pmap输入文件')
    parser.add_argument('-o', '--output', help='Excel输出文件 (例如: result.xlsx)')

    # 解析命令行参数
    args = parser.parse_args()

    # 读取pmap数据
    if args.input:
        try:
            with open(args.input, 'r') as f:
                pmap_content = f.readlines()
        except FileNotFoundError:
            print(f"错误：找不到文件 '{args.input}'")
            sys.exit(1)
    else:
        print("请输入pmap数据（输入结束后按Ctrl+D）：")
        pmap_content = sys.stdin.readlines()

    # 解析数据
    stats = parse_pmap_output(pmap_content)

    # 格式化输出（用于控制台）
    output_lines = format_output(stats)

    # 输出到Excel文件
    if args.output:
        if not args.output.lower().endswith(('.xlsx', '.xlsm')):
            print("警告：Excel文件扩展名应为.xlsx或.xlsm，已自动添加.xlsx")
            args.output += '.xlsx'
        write_to_excel(stats, args.output)

    # 控制台输出
    print("\n按Mapping类型统计的内存使用情况：")
    for line in output_lines:
        print(line)


if __name__ == "__main__":
    main()